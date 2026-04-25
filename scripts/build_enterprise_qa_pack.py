from pathlib import Path
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import tempfile

ROOT = Path(__file__).resolve().parents[1]
SHOT_DIR = ROOT / "docs" / "screenshots"
OUT = ROOT / "GNX_PRO_Enterprise_QA_Pack_WITH_SCREENSHOTS.xlsx"

TOOLS = [
    ("Dashboard", "dashboard.png"),
    ("AI Settings", "ai_settings.png"),
    ("Cloudinary", "cloudinary.png"),
    ("Repliz", "repliz.png"),
    ("Worker Control", "worker_control.png"),
    ("Monitoring", "monitoring.png"),
    ("About", "about.png"),
    ("Licensing", "licensing.png"),
]

header_fill = PatternFill("solid", fgColor="111111")
header_font = Font(bold=True, color="FFFFFF")
thin = Side(style="thin", color="666666")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

pass_fill = PatternFill("solid", fgColor="C6EFCE")
fail_fill = PatternFill("solid", fgColor="FFC7CE")
progress_fill = PatternFill("solid", fgColor="FFEB9C")
blocked_fill = PatternFill("solid", fgColor="D9D9D9")

def to_jpeg(src: Path, max_w=520, quality=70):
    im = PILImage.open(src).convert("RGB")
    w, h = im.size
    if w > max_w:
        nw = max_w
        nh = int(h * (nw / w))
        im = im.resize((nw, nh), PILImage.LANCZOS)
        w, h = im.size
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".jpg")
    im.save(tmp.name, "JPEG", quality=quality, optimize=True)
    return tmp.name, w, h

def add_img(ws, cell, path: Path):
    tmp, w, h = to_jpeg(path)
    img = XLImage(tmp)
    ws.add_image(img, cell)
    r = ws[cell].row
    ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or 15, h * 0.75)

wb = Workbook()

# ===== Tool Screenshots
ts = wb.active
ts.title = "Tool Screenshots"
ts["A1"] = "GNX PRO – Enterprise QA Pack (Screenshots)"
ts["A1"].font = Font(size=16, bold=True)
ts.merge_cells("A1:C1")
ts.column_dimensions["A"].width = 20
ts.column_dimensions["B"].width = 60
ts.column_dimensions["C"].width = 80

ts.append(["Tool/Page", "Purpose", "Screenshot"])
for c in range(1, 4):
    cell = ts.cell(2, c)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = border

row = 3
for tool, fname in TOOLS:
    ts[f"A{row}"] = tool
    ts[f"B{row}"] = "See Tool Guides / QA Checklist"
    for col in "AB":
        ts[f"{col}{row}"].border = border
        ts[f"{col}{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    p = SHOT_DIR / fname
    if p.exists():
        add_img(ts, f"C{row}", p)
        ts.row_dimensions[row].height = 260
    else:
        ts[f"C{row}"] = f"Missing: {fname} (capture and put into docs/screenshots)"
        ts[f"C{row}"].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    ts[f"C{row}"].border = border
    row += 1

# ===== QA Checklist
qc = wb.create_sheet("QA Checklist")
qc.freeze_panes = "A2"
headers = ["No","Module","Scenario","Expected","Plan","Severity","Status","Notes","Date"]
qc.append(headers)
for c in range(1, len(headers)+1):
    cell = qc.cell(1, c)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = border

for col,w in {"A":5,"B":16,"C":34,"D":30,"E":14,"F":10,"G":14,"H":24,"I":12}.items():
    qc.column_dimensions[col].width = w

rows = [
    (1,"Dashboard","Generate Offline Video","Job completes + output exists","FREE/BASIC/PRO","Medium","","",""),
    (2,"Dashboard","Generate YouTube URL","Download+process succeeds","PRO/TRIAL","High","","",""),
    (3,"AI","Hooks/Subtitles/Niche/Hashtags","AI outputs present","PRO/TRIAL","High","","",""),
    (4,"Cloudinary","Test + Upload","Cloudinary URL generated","BASIC/PRO","High","","",""),
    (5,"Repliz","Connect + accounts","Accounts loaded","BASIC/PRO","Critical","","",""),
    (6,"Repliz","Schedule after generate","Schedule created","BASIC/PRO","Critical","","",""),
    (7,"Worker","Start/Stop","Status updates correctly","BASIC/PRO","High","","",""),
    (8,"Monitoring","Jobs list","No permission errors","FREE/BASIC/PRO","Medium","","",""),
]
for r in rows:
    qc.append(list(r[:-1]) + [date.today().isoformat()])

dv_status = DataValidation(type="list", formula1='"PASS,FAIL,IN PROGRESS,BLOCKED"', allow_blank=True)
dv_sev = DataValidation(type="list", formula1='"Low,Medium,High,Critical"', allow_blank=True)
qc.add_data_validation(dv_status); qc.add_data_validation(dv_sev)
dv_status.add("G2:G200"); dv_sev.add("F2:F200")

qc.conditional_formatting.add("G2:G200", FormulaRule(formula=['$G2="PASS"'], fill=pass_fill))
qc.conditional_formatting.add("G2:G200", FormulaRule(formula=['$G2="FAIL"'], fill=fail_fill))
qc.conditional_formatting.add("G2:G200", FormulaRule(formula=['$G2="IN PROGRESS"'], fill=progress_fill))
qc.conditional_formatting.add("G2:G200", FormulaRule(formula=['$G2="BLOCKED"'], fill=blocked_fill))

for row in qc.iter_rows(min_row=2, max_row=qc.max_row, min_col=1, max_col=len(headers)):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="top")

# ===== Bug Log
bug = wb.create_sheet("Bug Log")
bug.freeze_panes = "A2"
bh = ["Bug ID","Title","Module","Severity","Status","Steps","Expected","Actual","Evidence"]
bug.append(bh)
for c in range(1, len(bh)+1):
    cell = bug.cell(1, c)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = border

wb.save(OUT)
print("Saved:", OUT)